
import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from fpdf import FPDF
import io
import base64

st.set_page_config(page_title="Taz-IT Invoice Generator", layout="centered")
st.title("📄 Taz-IT Invoice Generator")

# Invoice Meta
st.subheader("Invoice Details")
invoice_number = st.text_input("Invoice Number", value="1001")
invoice_date = st.date_input("Invoice Date", value=datetime.today())

# Client Info
st.subheader("Client Information")
client_name = st.text_input("Client Name", value="Bon Vibe")
client_address = st.text_input("Client Address", value="Kerkstraat 4, Aruba")
client_phone = st.text_input("Client Phone", value="297 746 3522")

# Line Items
st.subheader("Line Items")
item_df = st.data_editor(
    pd.DataFrame(columns=["Description", "Units", "Qty", "Rate (AWG)"]),
    num_rows="dynamic",
    use_container_width=True
)

def generate_pdf(invoice_number, invoice_date, client_name, client_address, client_phone, df):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(200, 10, "INVOICE", ln=True, align="R")

    # Company Info
    pdf.set_font("Arial", "", 10)
    pdf.cell(100, 8, "Taz-IT Solutions", ln=True)
    pdf.cell(100, 5, "Pos Chikito 99B", ln=True)
    pdf.cell(100, 5, "Oranjestad, Aruba", ln=True)
    pdf.cell(100, 5, "(+297) 699-7692 | jcroes@tazitsolution.com", ln=True)

    # Invoice and Client Info
    pdf.ln(5)
    pdf.cell(100, 6, f"Bill To: {client_name}", ln=False)
    pdf.cell(100, 6, f"Invoice #: {invoice_number}", ln=True)
    pdf.cell(100, 6, f"{client_address}", ln=False)
    pdf.cell(100, 6, f"Date: {invoice_date.strftime('%d-%b-%Y')}", ln=True)
    pdf.cell(100, 6, f"{client_phone}", ln=True)

    pdf.ln(5)
    pdf.set_fill_color(192, 0, 0)
    pdf.set_text_color(255)
    pdf.set_font("Arial", "B", 10)
    headers = ["Description", "Units", "Qty", "Rate", "Total"]
    col_widths = [60, 25, 20, 30, 30]
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 8, h, border=1, align='C', fill=True)
    pdf.ln()

    # Table content
    pdf.set_text_color(0)
    pdf.set_font("Arial", "", 10)
    subtotal = 0
    for _, row in df.iterrows():
        if pd.isna(row["Description"]) or row["Description"] == "":
            continue
        qty = float(row.get("Qty", 0))
        rate = float(row.get("Rate (AWG)", 0))
        total = qty * rate
        subtotal += total
        values = [row["Description"], row["Units"], qty, rate, total]
        for i, val in enumerate(values):
            pdf.cell(col_widths[i], 8, str(val), border=1)
        pdf.ln()

    # Totals
    tax = subtotal * 0.12
    grand_total = subtotal + tax
    pdf.ln(3)
    for label, value in [("Subtotal", subtotal), ("Tax (12%)", tax), ("Total", grand_total)]:
        pdf.cell(135, 8, label, align="R")
        pdf.cell(30, 8, f"{value:.2f} AWG", border=1, ln=True)

    # Footer
    pdf.ln(10)
    pdf.set_font("Arial", "I", 10)
    pdf.cell(0, 10, "Thank you for your business!", ln=True)
    pdf.cell(0, 5, "Payment due within 14 days.", ln=True)

    buffer = io.BytesIO()
    pdf.output(buffer)
    return buffer.getvalue()

if st.button("Generate Invoice"):
    valid_df = item_df.dropna(subset=["Description", "Qty", "Rate (AWG)"])
    valid_df = valid_df[valid_df["Description"] != ""]
    if valid_df.empty:
        st.warning("Please enter at least one line item.")
    else:
        pdf_bytes = generate_pdf(invoice_number, invoice_date, client_name, client_address, client_phone, valid_df)

        # Download Button
        st.download_button(
            label="📄 Download PDF Invoice",
            data=pdf_bytes,
            file_name=f"Invoice_{invoice_number}.pdf",
            mime="application/pdf"
        )

        # Preview
        st.subheader("🔍 PDF Preview")
        base64_pdf = base64.b64encode(pdf_bytes).decode('utf-8')
        pdf_display = f'<iframe src="data:application/pdf;base64,{base64_pdf}" width="100%" height="600px" type="application/pdf"></iframe>'
        st.markdown(pdf_display, unsafe_allow_html=True)
