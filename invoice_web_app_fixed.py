
import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import io

st.title("📄 Taz-IT Invoice Generator")

# Invoice Meta
st.subheader("Invoice Details")
invoice_number = st.text_input("Invoice Number", value="1001")
invoice_date = st.date_input("Invoice Date", value=datetime.today())

# Client Info
st.subheader("Client Information")
client_name = st.text_input("Client Name", value="Bon Vibe")
client_address = st.text_input("Client Address", value="Kerkstraat 4, Aruba")
client_phone = st.text_input("Client Phone", value="297 746 3522")

# Line Items
st.subheader("Line Items")
item_df = st.data_editor(
    pd.DataFrame(columns=["Description", "Units", "Qty", "Rate (AWG)"]),
    num_rows="dynamic",
    use_container_width=True
)

if st.button("Generate Invoice"):
    # Prepare workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # Styles
    bold = Font(bold=True)
    center = Alignment(horizontal='center')
    right = Alignment(horizontal='right')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # Header
    ws.merge_cells('D1:F1')
    ws['D1'] = "INVOICE"
    ws['D1'].font = Font(size=20, bold=True)
    ws['D1'].alignment = right

    # Company Info
    ws['A3'] = "Taz-IT Solutions"
    ws['A4'] = "Pos Chikito 99B"
    ws['A5'] = "Oranjestad, Aruba"
    ws['A6'] = "(+297) 699-7692"
    ws['A7'] = "jcroes@tazitsolution.com"

    # Invoice Info
    ws['E3'] = "Invoice No:"
    ws['F3'] = invoice_number
    ws['E4'] = "Date:"
    ws['F4'] = invoice_date.strftime("%d-%b-%Y")

    # Client Info
    ws['A9'] = "Bill To:"
    ws['A9'].font = bold
    ws['A10'] = client_name
    ws['A11'] = client_address
    ws['A12'] = client_phone

    # Table Headers
    headers = ["Description", "Units", "Qty", "Rate (AWG)", "Total (AWG)"]
    start_row = 14
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col)
        cell.value = header
        cell.font = bold
        cell.alignment = center
        cell.border = border

    # Line Items
    subtotal = 0
    valid_rows = item_df.dropna(subset=["Description", "Qty", "Rate (AWG)"])
    valid_rows = valid_rows[valid_rows["Description"] != ""]
    row_idx = start_row + 1

    for _, row in valid_rows.iterrows():
        try:
            qty = float(row.get("Qty", 0)) or 0
            rate = float(row.get("Rate (AWG)", 0)) or 0
            total = qty * rate
        except Exception as e:
            continue  # skip invalid rows

        subtotal += total
        values = [row["Description"], row["Units"], qty, rate, total]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col)
            c.value = val
            c.alignment = center if col != 1 else Alignment(horizontal='left')
            c.border = border
        row_idx += 1

    # Totals
    tax = subtotal * 0.12
    grand_total = subtotal + tax
    for i, (label, val) in enumerate([("Subtotal", subtotal), ("Tax (12%)", tax), ("Total", grand_total)]):
        ws.merge_cells(start_row=row_idx + i, start_column=1, end_row=row_idx + i, end_column=4)
        ws.cell(row=row_idx + i, column=1).value = label
        ws.cell(row=row_idx + i, column=1).alignment = right
        ws.cell(row=row_idx + i, column=1).font = bold
        ws.cell(row=row_idx + i, column=1).border = border

        ws.cell(row=row_idx + i, column=5).value = val
        ws.cell(row=row_idx + i, column=5).alignment = center
        ws.cell(row=row_idx + i, column=5).border = border

    # Footer
    ws[f'A{row_idx + 3}'] = "Thank you for your business!"
    ws[f'A{row_idx + 4}'] = "Payment due within 14 days."

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 20

    # Output download
    output = io.BytesIO()
    wb.save(output)
    st.download_button("📥 Download Invoice", output.getvalue(), file_name=f"Invoice_{invoice_number}.xlsx")
